"""
Excel Export für Produktfamilien

Exportiert Produktfamilien mit:
- Sheet 1: Übersicht (Frontend-Style)
- Sheet 2: Gemeinsame Codes (wenn vorhanden)
- Sheet 3+: Pro Gruppe mit ALLEN Levels, dedupliziert
"""

import tempfile
from datetime import datetime
from collections import defaultdict
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Helper: Pattern Berechnung
# ============================================================

def _compute_pattern_string(full_typecode: str) -> str:
    """
    Berechnet Pattern aus full_typecode.
    Beispiel: "BCC M313-0000-20" -> "3-4-4-2"
    """
    if not full_typecode:
        return ""
    
    parts = full_typecode.split('-')
    pattern_parts = []
    
    for part in parts:
        if ' ' in part:
            # Split bei Leerzeichen
            subparts = part.split()
            for subpart in subparts:
                if subpart.strip():
                    pattern_parts.append(str(len(subpart.strip())))
        else:
            if part.strip():
                pattern_parts.append(str(len(part.strip())))
    
    return '-'.join(pattern_parts)


# ============================================================
# Shared Codes Analysis
# ============================================================

def _analyze_shared_codes(cursor, family_id: int, groups: List[dict]) -> dict:
    """
    Findet Codes die in >1 Gruppe vorkommen.
    
    Returns:
        {
            'total': int,
            'by_level': {
                level: {
                    (code, name, label, label_en): {
                        'code': str,
                        'name': str,
                        'label': str,
                        'label_en': str,
                        'groups': [str]
                    }
                }
            }
        }
    """
    level_codes = defaultdict(set)  # level -> {(code, name, label, label_en, group_name)}
    
    for group in groups:
        gname = group['group_name']
        
        for pattern in group['patterns']:
            pstring = pattern['pattern_string'] if isinstance(pattern, dict) else pattern.pattern_string
            num_segs = len(pstring.split('-'))
            
            # Für Level 1, 2, 3... (nicht Level 0 = Familie)
            for level in range(1, num_segs):
                # Hole DISTINCT Codes die zur Familie gehören
                cursor.execute("""
                    SELECT DISTINCT n.code
                    FROM nodes n
                    JOIN node_paths p ON p.descendant_id = n.id
                    WHERE p.ancestor_id = ?
                    AND n.group_name = ? 
                    AND n.level = ?
                    AND n.code IS NOT NULL 
                    AND n.full_typecode IS NOT NULL
                """, (family_id, gname, level))
                
                for row in cursor.fetchall():
                    code = row['code']
                    
                    # Hole EINEN Beispiel-Node für Attribute
                    cursor.execute("""
                        SELECT n.id, n.name, n.full_typecode
                        FROM nodes n
                        JOIN node_paths p ON p.descendant_id = n.id
                        WHERE p.ancestor_id = ?
                        AND n.code = ? 
                        AND n.level = ? 
                        AND n.group_name = ?
                        AND n.full_typecode IS NOT NULL
                        LIMIT 1
                    """, (family_id, code, level, gname))
                    
                    node = cursor.fetchone()
                    if not node:
                        continue
                    
                    # Pattern check - muss zum aktuellen Schema passen
                    node_pattern = _compute_pattern_string(node['full_typecode'])
                    if node_pattern != pstring:
                        continue
                    
                    # Get labels
                    cursor.execute("""
                        SELECT label_de, label_en
                        FROM node_labels
                        WHERE node_id = ?
                        ORDER BY display_order
                    """, (node['id'],))
                    
                    labels = cursor.fetchall()
                    label_de = '\n\n'.join([l['label_de'] for l in labels if l['label_de']])
                    label_en = '\n\n'.join([l['label_en'] for l in labels if l['label_en']])
                    name = node['name'] or ''
                    
                    level_codes[level].add((code, name, label_de, label_en, gname))
    
    # Find shared codes
    shared_by_level = {}
    total = 0
    
    for level, codes_set in level_codes.items():
        code_groups = defaultdict(list)
        
        for code, name, label_de, label_en, gname in codes_set:
            key = (code, name, label_de, label_en)
            code_groups[key].append(gname)
        
        shared_by_level[level] = {}
        for key, gnames in code_groups.items():
            unique_groups = list(set(gnames))
            if len(unique_groups) > 1:
                code, name, label_de, label_en = key
                shared_by_level[level][key] = {
                    'code': code,
                    'name': name,
                    'label': label_de,
                    'label_en': label_en,
                    'groups': sorted(unique_groups)
                }
                total += 1
    
    return {'total': total, 'by_level': shared_by_level}


# ============================================================
# Overview Sheet
# ============================================================

def _create_overview_sheet(ws, family_code: str, family_label: str, groups: List[dict]):
    """Erstellt Übersicht-Sheet (Frontend-Style)"""
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:E{current_row}')
    title = ws.cell(row=current_row, column=1, value=f"Produktfamilie: {family_code}")
    title.font = Font(bold=True, size=16)
    title.alignment = Alignment(horizontal="center")
    current_row += 1
    
    if family_label:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        subtitle = ws.cell(row=current_row, column=1, value=family_label)
        subtitle.font = Font(size=12)
        subtitle.alignment = Alignment(horizontal="center")
        current_row += 2
    else:
        current_row += 1
    
    # Groups
    for group in groups:
        gname = group['group_name']
        patterns = group['patterns']
        
        ws.merge_cells(f'A{current_row}:E{current_row}')
        gh = ws.cell(row=current_row, column=1, value=f"Gruppe: {gname}")
        gh.font = Font(bold=True, size=12, color="FFFFFF")
        gh.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        current_row += 1
        
        for pattern in patterns:
            if isinstance(pattern, dict):
                pstring = pattern['pattern_string']
                examples = pattern['segment_examples']
                count = pattern['count']
            else:
                pstring = pattern.pattern_string
                examples = pattern.segment_examples
                count = pattern.count
            
            example_str = '-'.join(examples)
            ws.cell(row=current_row, column=1, value=f"Schema: {pstring}").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=f"Beispiel: {example_str}")
            ws.cell(row=current_row, column=3, value=f"Anzahl: {count}")
            current_row += 1
        
        current_row += 1


# ============================================================
# Shared Codes Sheet
# ============================================================

def _create_shared_codes_sheet(ws, shared_data: dict):
    """Erstellt Sheet mit gemeinsamen Codes"""
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Title
    ws.cell(row=current_row, column=1, value="Gemeinsame Codes über mehrere Gruppen").font = Font(bold=True, size=14)
    current_row += 2
    
    # Pro Level
    for level in sorted(shared_data['by_level'].keys()):
        codes_dict = shared_data['by_level'][level]
        
        if not codes_dict:
            continue
        
        # Level Header
        ws.cell(row=current_row, column=1, value=f"Level {level} ({len(codes_dict)} Codes)").font = Font(bold=True, size=11)
        current_row += 1
        
        # Table Header
        headers = ["Code", "Name", "Label (DE)", "Label (EN)", "Gruppen"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.border = border
        current_row += 1
        
        # Data
        for key in sorted(codes_dict.keys(), key=lambda x: x[0]):
            data = codes_dict[key]
            row_data = [
                data['code'],
                data['name'],
                data['label'][:100] + '...' if len(data['label']) > 100 else data['label'],
                data['label_en'][:100] + '...' if len(data['label_en']) > 100 else data['label_en'],
                ', '.join(data['groups'])
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            current_row += 1
        
        current_row += 2


# ============================================================
# Group Sheet - ALLE LEVELS
# ============================================================

def _create_group_sheet(ws, cursor, family_id: int, family_code: str, group: dict, shared_codes: dict):
    """
    Erstellt Sheet für eine Gruppe - ALLE LEVELS.
    
    WICHTIG:
    - Iteriert über ALLE Levels (1, 2, 3...) nicht nur das letzte
    - Dedupliziert nach (code, name, label, label_en)
    - Pfad-Kontext NUR bei Duplikaten
    """
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    gname = group['group_name']
    patterns = group['patterns']
    
    current_row = 1
    
    # Group Title
    ws.cell(row=current_row, column=1, value=f"Gruppe: {gname}").font = Font(bold=True, size=14)
    current_row += 2
    
    # Pro Schema
    for pattern in patterns:
        if isinstance(pattern, dict):
            pstring = pattern['pattern_string']
            seg_names = pattern['segment_names']
            count = pattern['count']
        else:
            pstring = pattern.pattern_string
            seg_names = pattern.segment_names
            count = pattern.count
        
        num_segs = len(pstring.split('-'))
        
        # Schema Header
        ws.cell(row=current_row, column=1, value=f"Schema: {pstring} ({count} Codes)").font = Font(bold=True, size=11)
        current_row += 2
        
        # WICHTIG: Für JEDES Level (1, 2, 3...) eine Tabelle
        for level in range(1, num_segs):
            level_name = seg_names[level] if level < len(seg_names) and seg_names[level] else f"Level {level}"
            
            # Hole ALLE Nodes auf diesem Level, die zur Familie gehören und zum Pattern passen
            cursor.execute("""
                SELECT DISTINCT n.id, n.code, n.name, n.full_typecode
                FROM nodes n
                JOIN node_paths p ON p.descendant_id = n.id
                WHERE p.ancestor_id = ?
                AND n.level = ? 
                AND n.group_name = ?
                AND n.code IS NOT NULL 
                AND n.full_typecode IS NOT NULL
            """, (family_id, level, gname))
            
            all_nodes = cursor.fetchall()
            if not all_nodes:
                continue
            
            # Dedupliziere nach (code, name, label, label_en)
            codes_dict = {}  # (code, name, label, label_en) -> set(paths)
            
            for node in all_nodes:
                # Pattern check - muss zum aktuellen Schema passen
                node_pattern = _compute_pattern_string(node['full_typecode'])
                if node_pattern != pstring:
                    continue
                
                code = node['code']
                name = node['name'] or ''
                node_id = node['id']
                
                # Get labels
                cursor.execute("""
                    SELECT label_de, label_en
                    FROM node_labels
                    WHERE node_id = ?
                    ORDER BY display_order
                """, (node_id,))
                
                labels = cursor.fetchall()
                label_de = '\n\n'.join([l['label_de'] for l in labels if l['label_de']])
                label_en = '\n\n'.join([l['label_en'] for l in labels if l['label_en']])
                
                key = (code, name, label_de, label_en)
                
                # Skip if shared
                if level in shared_codes['by_level'] and key in shared_codes['by_level'][level]:
                    continue
                
                # Get path
                cursor.execute("""
                    SELECT n2.code
                    FROM node_paths p
                    JOIN nodes n2 ON p.ancestor_id = n2.id
                    WHERE p.descendant_id = ? AND p.ancestor_id != p.descendant_id
                    ORDER BY n2.level
                """, (node_id,))
                
                path_codes = [r['code'] for r in cursor.fetchall() if r['code']]
                path_str = ' → '.join(path_codes)
                
                if key not in codes_dict:
                    codes_dict[key] = set()
                if path_str:
                    codes_dict[key].add(path_str)
            
            if not codes_dict:
                continue
            
            # Level Header
            ws.cell(row=current_row, column=1, value=f"{level_name} ({len(codes_dict)} Varianten)").font = Font(bold=True, size=10)
            current_row += 1
            
            # Table Header
            headers = ["Pfad", "Code", "Name", "Label (DE)", "Label (EN)"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.border = border
            current_row += 1
            
            # Data
            for (code, name, label_de, label_en), paths in sorted(codes_dict.items(), key=lambda x: x[0][0]):
                # Pfad NUR wenn mehrere (= Duplikate)
                if len(paths) > 1:
                    for path in sorted(paths):
                        row_data = [
                            path, code, name,
                            label_de[:100] + '...' if len(label_de) > 100 else label_de,
                            label_en[:100] + '...' if len(label_en) > 100 else label_en
                        ]
                        
                        for col, val in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col, value=val)
                            cell.border = border
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
                            if col == 1:  # Pfad highlight
                                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                                cell.font = Font(size=8, italic=True)
                        current_row += 1
                else:
                    # Kein Pfad (einzigartig)
                    row_data = [
                        '', code, name,
                        label_de[:100] + '...' if len(label_de) > 100 else label_de,
                        label_en[:100] + '...' if len(label_en) > 100 else label_en
                    ]
                    
                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=val)
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                    current_row += 1
            
            current_row += 2  # Gap between levels
        
        current_row += 1  # Gap between schemas


# ============================================================
# Main Export Function
# ============================================================

def export_family_to_excel(cursor, family_code: str) -> tuple:
    """
    Exportiert Produktfamilie als Excel.
    
    Returns:
        (temp_file_path, filename)
    
    Raises:
        ValueError: Wenn Familie nicht gefunden oder keine Daten
    """
    # 1. Get family
    cursor.execute("SELECT id, code, label FROM nodes WHERE code = ? AND level = 0", (family_code,))
    family = cursor.fetchone()
    if not family:
        raise ValueError(f"Familie '{family_code}' nicht gefunden")
    
    family_id = family['id']
    family_label = family['label'] or family_code
    
    # 2. Get groups
    cursor.execute("""
        SELECT DISTINCT n.group_name
        FROM nodes n
        JOIN node_paths p ON p.descendant_id = n.id
        WHERE p.ancestor_id = ? AND n.group_name IS NOT NULL
        ORDER BY n.group_name
    """, (family_id,))
    
    group_names = [row[0] for row in cursor.fetchall()]
    if not group_names:
        raise ValueError("Keine Gruppen gefunden")
    
    # 3. Analyze schemas per group
    # Import hier um zirkuläre Imports zu vermeiden
    from api import _analyze_schemas_for_group
    
    groups = []
    for gname in group_names:
        patterns = _analyze_schemas_for_group(cursor, family_id, family_code, gname)
        if patterns:
            groups.append({'group_name': gname, 'patterns': patterns})
    
    if not groups:
        raise ValueError("Keine exportierbaren Daten")
    
    # 4. Analyze shared codes
    shared_codes = _analyze_shared_codes(cursor, family_id, groups)
    
    # 5. Create Workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Overview
    ws_overview = wb.create_sheet(title=f"Übersicht {family_code}")
    _create_overview_sheet(ws_overview, family_code, family_label, groups)
    
    # Sheet 2: Shared Codes (wenn vorhanden)
    if shared_codes['total'] > 0:
        ws_shared = wb.create_sheet(title="Gemeinsame Codes")
        _create_shared_codes_sheet(ws_shared, shared_codes)
    
    # Sheets 3+: Pro Gruppe
    for group in groups:
        gname = group['group_name'][:31].replace('/', '-').replace('\\', '-').replace(':', '-')
        ws_group = wb.create_sheet(title=gname)
        _create_group_sheet(ws_group, cursor, family_id, family_code, group, shared_codes)
    
    # 6. Save
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp.name)
    temp.close()
    
    filename = f"{family_code}_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return (temp.name, filename)
